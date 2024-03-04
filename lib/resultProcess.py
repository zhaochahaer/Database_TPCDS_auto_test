
import os
import re
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from lib.Logger import logger
from openpyxl.styles import Alignment, Border, Side
import pandas as pd
import codecs
from lib.mppInfo import Userinof
from lib.mppQbmconfig import *

class ResultProcess:
    def __init__(self):
        self.date = datetime.now().strftime('%Y%m%d')
        self.tpcds_log_path = os.path.join(os.getcwd(), 'result', 'tpcds')
        self.tpcds_excel_path = os.path.join(os.getcwd(),'html','tpcds')
        self.tpcds_moban=os.path.join(os.getcwd(),'doc','performance_tpcds_moban.xlsx')
        self.tpcds_all_moban=os.path.join(os.getcwd(),'doc','performance_tpcds_all_moban.xlsx')

    def read_tpcds_log_list(self):
        # log_tpcds_list=[]
        filename=os.path.join(self.tpcds_log_path,'log_tpcds_list')
        
        with open(filename,'r+') as f:
            content = f.read().split()
            # log_tpcds_list.append(content)

        return content
    
    def tpcds_result(self, log_file):
        tpcds_load_table = []
        tpcds_generate = []
        tpcds_select_times = []
        tpcds_analyze_times = []

        with open(log_file, 'r', encoding="utf-8") as f:
            all_content = f.read()

            lines_load = re.findall(r'Data Loads\n.*?\n(.*?)\n\(24 rows\)', all_content, re.S)
            for line in lines_load:
                line1 = line.split('\n')
                for t1 in line1:
                    tpcds_load_table.append(t1.split('\t'))

            lines_generate = re.findall(r'Generate Data\n.*?\n(.*?)\n\(3 rows\)', all_content, re.S)
            for line_g in lines_generate:
                line_g1 = line_g.split('\n')
                for t2 in line_g1:
                    tpcds_generate.append(t2.split('\t'))

            lines_queries = re.findall(r'Queries\n.*?\n(.*?)\n\(99 rows\)', all_content, re.S)
            for line_q in lines_queries:
                line_q1 = line_q.split('\n')
                for t3 in line_q1:
                    tpcds_select_times.append(t3.split('\t'))

            lines_analyze = re.findall(r'Analyze\n.*?\n(.*?)\n\(1 row\)', all_content, re.S)
            for line_a in lines_analyze:
                line_a1 = line_a.split('\n')
                for t4 in line_a1:
                    tpcds_analyze_times.append(t4.split('\t'))
   
        return tpcds_load_table, tpcds_generate, tpcds_select_times, tpcds_analyze_times

    def performance_tpcds_excel(self, log_files):
        date_folder = os.path.join(self.tpcds_excel_path, self.date)
        os.makedirs(date_folder, exist_ok=True)

        excel_file = os.path.join(date_folder, 'performance_tpcds_{}.xlsx'.format(self.date))

        shutil.copyfile(self.tpcds_moban, excel_file)

        workbook = load_workbook(excel_file)
        sheet = workbook.active
        start_row = 1
        start_column = 4
        avg_column = 4
        
        sheet.row_dimensions[start_row].height = 44
        # print(f"????{log_files}")
        for log_file in log_files:
            # print(f"!!!!!{log_file}")
            tpcds_load_table, tpcds_generate, tpcds_select_times, tpcds_analyze_times = self.tpcds_result(log_file)

            cell_identifier = chr(start_column - 1 + 65) + str(start_row)

            # 1. 将日期和耗时信息写入第1行、第4列
            text = '日期：{}\nGenerate Data耗时(s):{}\nAnalyze耗时(s):{}'.format(
                self.date, round(float(tpcds_generate[1][1]), 2), round(float(tpcds_analyze_times[1][1]), 2)
            )
            sheet[cell_identifier] = text
            alignment = Alignment(wrap_text=True)
            sheet[cell_identifier].alignment = alignment

            # 2. 将标题写入第2行、第4列
            sheet.cell(row=start_row + 1, column=start_column).value = '耗时（秒）'

            # 3. 将tpcds_load_table数据写入第3-26行、第4列
            for i, row_data in enumerate(tpcds_load_table[1:], start=start_row + 2):
                sheet.cell(row=i, column=start_column).value = float(row_data[2])

            # 4. 将第3-26行、第4列的数据求和，放在第27行、第4列
            # sum_formula = '=SUM({}{}:{}{})'.format(
            #     get_column_letter(start_column), start_row + 2,
            #     get_column_letter(start_column), start_row + 25
            # )
            sum_formula1 = sum(float(row_data[2]) for row_data in tpcds_load_table[1:])
            sheet.cell(row=start_row + 26, column=start_column).value = sum_formula1

            # 5.1. 将标题写入第29行、第4列
            sheet.cell(row=start_row + 28, column=start_column).value = 'duration'

            # 5. 将tpcds_select_times数据写入第30-128行、第4列
            for i, row_data in enumerate(tpcds_select_times[1:], start=start_row + 29):
                sheet.cell(row=i, column=start_column).value = float(row_data[2])

            # 6. 将第30-128行、第4列的数据求和，放在第129行、第4列
            # sum_formula = '=SUM({}{}:{}{})'.format(
            #     get_column_letter(start_column), start_row + 29,
            #     get_column_letter(start_column), start_row + 127
            # )
            sum_formula2 = sum(float(row_data[2]) for row_data in tpcds_select_times[1:])
            sheet.cell(row=start_row + 128, column=start_column).value = sum_formula2

            # 7. 将第129行、第4列和第27行、第4列的数据求和，放在第130行、第4列
            # sum_formula = '={}27 + {}129'.format(get_column_letter(start_column), get_column_letter(start_column))
            # sheet.cell(row=start_row + 129, column=start_column).value = sum_formula
            sum_formula = sum_formula1 + sum_formula2
            sheet.cell(row=start_row + 129, column=start_column).value = sum_formula

            # 设置列宽为28
            sheet.column_dimensions[get_column_letter(start_column)].width = 28

            # 设置边框样式
            border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in sheet.iter_rows(min_row=start_row, max_row=start_row + 129, min_col=start_column, max_col=start_column):
                for cell in row:
                    cell.border = border_style

            start_column+=1


        avg_row = start_row + 2
        sheet.cell(row=start_row, column=start_column).value = '平均值'
        sheet.cell(row=start_row + 1, column=start_column).value = '耗时（秒）'


        # 计算tpcds_load_table每行的平均值
        for row in sheet.iter_rows(min_row=start_row + 2, max_row=start_row + 26, min_col=avg_column, max_col=start_column - 1):
            # avg_formula = '=AVERAGE({}:{})'.format(row[0].coordinate, row[-1].coordinate)
            # sheet.cell(row=avg_row, column=start_column).value = avg_formula
            # sheet.cell(row=avg_row, column=start_column).number_format = '0.00'
            # avg_row += 1
            values = [cell.value for cell in row]
            avg_value1 = sum(float(value) for value in values) / len(values)
            sheet.cell(row=avg_row, column=start_column).value = avg_value1
            sheet.cell(row=avg_row, column=start_column).number_format = '0.00'
            avg_row += 1

        # 计算tpcds_select_times每行的平均值
        for row_cells in sheet.iter_rows(min_row=start_row + 29, max_row=start_row + 129, min_col=avg_column, max_col=start_column - 1):
            # avg_formula = '=AVERAGE({}:{})'.format(row_cells[0].coordinate, row_cells[-1].coordinate)
            # sheet.cell(row=row_cells[0].row, column=start_column).value = avg_formula
            # sheet.cell(row=row_cells[0].row, column=start_column).number_format = '0.00'
            values = [cell.value for cell in row_cells]
            avg_value2 = sum(float(value) for value in values) / len(values)
            sheet.cell(row=row_cells[0].row, column=start_column).value = avg_value2
            sheet.cell(row=row_cells[0].row, column=start_column).number_format = '0.00'

        
            # 设置边框样式
        for row in sheet.iter_rows(min_row=start_row, max_row=start_row + 129, min_col=avg_column, max_col=start_column):
            for cell in row:
                cell.border = border_style

        # 设置列宽为28
        sheet.column_dimensions[get_column_letter(start_column)].width = 28
        
        # 保存 Excel 文件
        workbook.save(excel_file)
        logger.info("make Excel complete!")
        self.save_as_mht(excel_file)
        logger.info(" Excel to HTML completed!")
        
        return excel_file

    #EXCEL to HTML   
    def save_as_mht(self,excel_file):
        xd = pd.ExcelFile(excel_file)
        df = xd.parse()
        df = df.fillna('')
        html_str = df.to_html(header = True,index = False, col_space=100)
        style = '''
        <style>
            table{
                border-spacing:0;  
            }
            th{
                text-align:center;vertical-align:middle;
            }
            td{
                text-align:center;vertical-align:middle;
            }
        </style>
        '''
        html_str = style + html_str
        mht_file = os.path.splitext(excel_file)[0] + ".html"
        with codecs.open(mht_file,'w','utf-8') as html_file:
            html_file.write(html_str)

        return mht_file
    
    #performance_tpcds_all.xlsx
    def performance_tpcds_excel_all(self, start_test_time,excel_file):
        logger.info("performance_tpcds_excel_all")
        excel_file_all = os.path.join(self.tpcds_excel_path, "performance_tpcds_all.xlsx")

        if not os.path.exists(excel_file_all):
            shutil.copyfile(self.tpcds_all_moban, excel_file_all)

        wb = load_workbook(excel_file)

        # 获取最后一列的数据
        avg_value_load = wb["Sheet1"].cell(row=27, column=wb["Sheet1"].max_column).value
        avg_value_select = wb["Sheet1"].cell(row=129, column=wb["Sheet1"].max_column).value
        avg_all = wb["Sheet1"].cell(row=130, column=wb["Sheet1"].max_column).value

        # 打开performance_tpcds_all.xlsx
        excel_file_all = os.path.join(self.tpcds_excel_path, "performance_tpcds_all.xlsx")
        wb_all = load_workbook(excel_file_all)

        # 针对TotalTime
        ws_total_time_all = wb_all["TotalTime"]
        row_count = ws_total_time_all.max_row + 1
        ws_total_time_all.cell(row=row_count, column=1).value = self.date
        ws_total_time_all.cell(row=row_count, column=2).value = start_test_time
        ws_total_time_all.cell(row=row_count, column=3).value = Userinof.tpcds_num
        ws_total_time_all.cell(row=row_count, column=4).value = avg_value_load
        ws_total_time_all.cell(row=row_count, column=5).value = avg_value_select
        ws_total_time_all.cell(row=row_count, column=6).value = avg_all
        for i in range(4, 7):
            ws_total_time_all.cell(row=row_count, column=i).number_format = '0.00'

        # 设置边框样式为所有边框
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws_total_time_all.iter_rows(min_row=row_count, max_row=row_count, min_col=1, max_col=6):
            for cell in row:
                cell.border = border

        # 针对QueryAll
        ws_query_all = wb_all["QueryAll"]
        logger.info(f"ws_query_all.max_column:{ws_query_all.max_column}")
        if ws_query_all.max_column <= 4:
            column_count = ws_query_all.max_column
            ws_query_all.insert_cols(column_count)
            ws_query_all.cell(row=1, column=column_count).value = self.date
            # ws_query_all.cell(row=row, column=column_count).border = border
            for row in range(2, 131):
                value = wb["Sheet1"].cell(row=row, column=wb["Sheet1"].max_column).value
                ws_query_all.cell(row=row, column=column_count).value = value
                ws_query_all.cell(row=row, column=column_count).border = border
        else:
            column_count = ws_query_all.max_column
            ws_query_all.insert_cols(column_count)
            ws_query_all.cell(row=1, column=column_count).value = self.date
            ws_query_all.cell(row=2, column=column_count).value = "耗时(秒)"
            for row in range(3, 131):
                current_value = wb["Sheet1"].cell(row=row, column=wb["Sheet1"].max_column).value
                ws_query_all.cell(row=row, column=column_count).value = current_value
                previous_value = ws_query_all.cell(row=row, column=column_count-1).value
                if previous_value is not None and current_value is not None:
                    percentage_change = (float(current_value) - float(previous_value)) / float(previous_value)
                    ws_query_all.cell(row=row, column=column_count+1).value = percentage_change
                    ws_query_all.cell(row=row, column=column_count+1).number_format = '0.00'
                else:
                    ws_query_all.cell(row=row, column=column_count+1).value = None
                ws_query_all.cell(row=row, column=column_count).border = border

        wb_all.save(excel_file_all)

        return excel_file_all
